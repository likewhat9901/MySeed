"""Read an uploaded xlsx and produce a compact per-sheet summary that an LLM can reason over."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MAX_SAMPLE_ROWS = 8
MAX_DISTINCT_SAMPLES = 5
MAX_HEADER_LEN = 60
MAX_ROWS_SCAN = 5000  # safety cap when profiling columns


@dataclass
class ColumnProfile:
    letter: str
    header: str
    inferred_type: str  # number | string | date | boolean | mixed | empty
    non_null: int
    distinct_sample: list[Any]
    numeric_stats: dict[str, float] | None


@dataclass
class SheetSummary:
    name: str
    n_rows: int
    n_cols: int
    header_row: int  # 1-indexed Excel row
    headers: list[str]
    sample_rows: list[list[Any]]
    columns: list[ColumnProfile]
    total_range: str  # e.g., "A1:E120"


@dataclass
class WorkbookSummary:
    sheets: list[SheetSummary]


def _infer_type(values: list[Any]) -> str:
    types: set[str] = set()
    has_value = False
    for v in values:
        if v is None or v == "":
            continue
        has_value = True
        if isinstance(v, bool):
            types.add("boolean")
        elif isinstance(v, (int, float)):
            types.add("number")
        elif hasattr(v, "isoformat"):
            types.add("date")
        else:
            types.add("string")
    if not has_value:
        return "empty"
    if len(types) == 1:
        return next(iter(types))
    return "mixed"


def _detect_header_row(rows: list[list[Any]]) -> int:
    """Pick first row (within first 5) that is mostly non-empty and mostly text."""
    for i, row in enumerate(rows[:5]):
        non_null = [v for v in row if v not in (None, "")]
        if len(non_null) >= max(2, len(row) // 2):
            text_cells = sum(1 for v in non_null if isinstance(v, str))
            if non_null and text_cells / len(non_null) >= 0.6:
                return i + 1
    return 1


def _jsonable_one(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float, str)):
        return v
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


def _jsonable_row(row: list[Any]) -> list[Any]:
    return [_jsonable_one(v) for v in row]


def summarize_workbook(content: bytes) -> WorkbookSummary:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheets: list[SheetSummary] = []

    try:
        for name in wb.sheetnames:
            ws = wb[name]
            all_rows: list[list[Any]] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx >= MAX_ROWS_SCAN:
                    break
                all_rows.append(list(row))

            if not all_rows:
                sheets.append(
                    SheetSummary(
                        name=name,
                        n_rows=0,
                        n_cols=0,
                        header_row=1,
                        headers=[],
                        sample_rows=[],
                        columns=[],
                        total_range="A1:A1",
                    )
                )
                continue

            n_cols = max((len(r) for r in all_rows), default=0)
            header_row = _detect_header_row(all_rows)
            headers_raw = all_rows[header_row - 1] if header_row - 1 < len(all_rows) else []
            headers: list[str] = []
            for i in range(n_cols):
                h = headers_raw[i] if i < len(headers_raw) else None
                headers.append(str(h)[:MAX_HEADER_LEN] if h is not None else f"col{i + 1}")

            data_rows = all_rows[header_row:]
            sample = [
                _jsonable_row(list(r) + [None] * (n_cols - len(r)))
                for r in data_rows[:MAX_SAMPLE_ROWS]
            ]

            columns: list[ColumnProfile] = []
            for i in range(n_cols):
                values = [r[i] if i < len(r) else None for r in data_rows]
                inferred = _infer_type(values)
                non_null = sum(1 for v in values if v not in (None, ""))

                distinct: list[Any] = []
                seen: set[str] = set()
                for v in values:
                    if v in (None, ""):
                        continue
                    key = repr(v)
                    if key in seen:
                        continue
                    seen.add(key)
                    distinct.append(_jsonable_one(v))
                    if len(distinct) >= MAX_DISTINCT_SAMPLES:
                        break

                stats: dict[str, float] | None = None
                if inferred == "number":
                    nums = [
                        float(v)
                        for v in values
                        if isinstance(v, (int, float)) and not isinstance(v, bool)
                    ]
                    if nums:
                        stats = {
                            "min": min(nums),
                            "max": max(nums),
                            "sum": sum(nums),
                            "avg": sum(nums) / len(nums),
                            "count": float(len(nums)),
                        }

                columns.append(
                    ColumnProfile(
                        letter=get_column_letter(i + 1),
                        header=headers[i] if i < len(headers) else f"col{i + 1}",
                        inferred_type=inferred,
                        non_null=non_null,
                        distinct_sample=distinct,
                        numeric_stats=stats,
                    )
                )

            last_data_row = header_row + len(data_rows)
            last_col_letter = get_column_letter(max(1, n_cols))
            sheets.append(
                SheetSummary(
                    name=name,
                    n_rows=last_data_row,
                    n_cols=n_cols,
                    header_row=header_row,
                    headers=headers,
                    sample_rows=sample,
                    columns=columns,
                    total_range=f"A1:{last_col_letter}{last_data_row}",
                )
            )
    finally:
        wb.close()

    return WorkbookSummary(sheets=sheets)


def serialize_for_llm(summary: WorkbookSummary) -> dict[str, Any]:
    return {
        "sheets": [
            {
                "name": s.name,
                "header_row": s.header_row,
                "n_rows": s.n_rows,
                "n_cols": s.n_cols,
                "total_range": s.total_range,
                "headers": s.headers,
                "sample_rows": s.sample_rows,
                "columns": [
                    {
                        "letter": c.letter,
                        "header": c.header,
                        "type": c.inferred_type,
                        "non_null": c.non_null,
                        "distinct_sample": c.distinct_sample,
                        "numeric_stats": c.numeric_stats,
                    }
                    for c in s.columns
                ],
            }
            for s in summary.sheets
        ]
    }


def serialize_summary(summary: WorkbookSummary) -> dict[str, Any]:
    """Same as serialize_for_llm; kept as a public alias for the API response."""
    return serialize_for_llm(summary)


def find_sheet(summary: WorkbookSummary, name: str) -> SheetSummary | None:
    for s in summary.sheets:
        if s.name == name:
            return s
    return None
