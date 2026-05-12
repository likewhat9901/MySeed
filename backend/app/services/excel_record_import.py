"""Parse spreadsheet rows into `tb_record.data`-shaped payloads for ledger import."""

from __future__ import annotations

import io
import math
import re
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

MAX_IMPORT_ROWS = 5000

# Logical roles → 헤더 셀에 적힌 문자열(column_map 값)
RECORD_ROLES_DATE = "date"
RECORD_ROLES_AMOUNT = "amount"
RECORD_ROLES_TITLE = "title"
RECORD_ROLES_MEMO = "memo"
RECORD_ROLES_CATEGORY = "category"


def _norm_header(v: Any) -> str:
    return str(v).strip() if v not in (None, "") else ""


def _detect_header_row(all_rows: list[list[Any]]) -> int:
    for i, row in enumerate(all_rows[:5]):
        non_null = [v for v in row if v not in (None, "")]
        if len(non_null) >= max(2, len(row) // 2):
            text_cells = sum(1 for v in non_null if isinstance(v, str))
            if non_null and text_cells / len(non_null) >= 0.6:
                return i + 1
    return 1


def _resolve_col_index(headers: list[str], label: str | None) -> int | None:
    if label is None or not str(label).strip():
        return None
    want = _norm_header(label)
    low = want.lower()
    for i, h in enumerate(headers):
        hn = _norm_header(h)
        if hn == want or hn.lower() == low:
            return i
    return None


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None:
        return None
    return row[idx] if idx < len(row) else None


def _jsonable_scalar(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, float) and math.isnan(v):  # type: ignore[arg-type]
        return None
    if isinstance(v, (int, float, str)):
        return v
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


def _coerce_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return float(int(v))
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):  # type: ignore[arg-type]
            return None
        return float(v)
    if isinstance(v, Decimal):
        try:
            return float(v)
        except (ValueError, InvalidOperation):
            return None
    s = str(v).strip().replace(",", "").replace("₩", "").replace("원", "").strip()
    if not s or s.lower() == "nan":
        return None
    try:
        return float(Decimal(s))
    except InvalidOperation:
        pass
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if m:
        try:
            return float(m.group())
        except ValueError:
            return None
    return None


def read_sheet_tabular(content: bytes, sheet_name: str | None = None) -> tuple[str, int, list[str], list[list[Any]]]:
    """
    Returns `(used_sheet_name, header_row_1based, headers as strings, raw data rows)`
    각 data row는 시트 행 하나(패딩 없음 가능).
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        names = wb.sheetnames
        if not names:
            return "", 1, [], []
        resolved = sheet_name.strip() if sheet_name and sheet_name.strip() else None
        if resolved:
            ws = wb[resolved] if resolved in names else None
            if ws is None:
                return resolved, 1, [], []
        else:
            ws = wb[names[0]]
            resolved = names[0]

        all_rows: list[list[Any]] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx >= MAX_IMPORT_ROWS:
                break
            all_rows.append(list(row))

        if not all_rows or not ws:
            return resolved, 1, [], []

        n_cols = max((len(r) for r in all_rows), default=0)
        header_row = _detect_header_row(all_rows)
        headers_raw = all_rows[header_row - 1] if header_row - 1 < len(all_rows) else []
        headers: list[str] = []
        for i in range(n_cols):
            h = headers_raw[i] if i < len(headers_raw) else None
            headers.append(str(h).strip() if h not in (None, "") else f"column_{i + 1}")

        data_rows = all_rows[header_row:]
        return resolved, header_row, headers, data_rows
    finally:
        wb.close()


def build_record_rows(
    *,
    led_id: UUID,
    data_type: str,
    sheet: str,
    header_row_1based: int,
    headers: list[str],
    data_rows: list[list[Any]],
    column_map: dict[str, str],
    file_id: UUID | None,
    skip_empty_amount: bool = True,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    `column_map` 키: date | amount | title | memo | category (값은 엑셀 헤더 텍스트).
    반환: insert용 dict 목록 및 경고 문자열 목록.
    """
    warnings: list[str] = []

    ix_date = _resolve_col_index(headers, column_map.get(RECORD_ROLES_DATE))
    ix_amount = _resolve_col_index(headers, column_map.get(RECORD_ROLES_AMOUNT))
    ix_title = _resolve_col_index(headers, column_map.get(RECORD_ROLES_TITLE))
    ix_memo = _resolve_col_index(headers, column_map.get(RECORD_ROLES_MEMO))
    ix_cat = _resolve_col_index(headers, column_map.get(RECORD_ROLES_CATEGORY))

    required = RECORD_ROLES_DATE, RECORD_ROLES_AMOUNT
    for role in required:
        if not str(column_map.get(role, "")).strip():
            warnings.append(f"column_map에 '{role}' 헤더가 비어 있습니다.")
    if ix_date is None and column_map.get(RECORD_ROLES_DATE):
        warnings.append(f"'{column_map.get(RECORD_ROLES_DATE)}' 헤더를 시트에서 찾지 못했습니다.")
    if ix_amount is None and column_map.get(RECORD_ROLES_AMOUNT):
        warnings.append(f"'{column_map.get(RECORD_ROLES_AMOUNT)}' 헤더를 시트에서 찾지 못했습니다.")
    if ix_date is None or ix_amount is None:
        return [], warnings

    out: list[dict[str, Any]] = []
    for offset, raw in enumerate(data_rows):
        excel_row_no = header_row_1based + 1 + offset
        padded = raw + [None] * max(0, len(headers) - len(raw))

        amt = _coerce_amount(_cell(padded, ix_amount))
        if skip_empty_amount and (amt is None):
            continue
        if amt is None and not skip_empty_amount:
            amt = 0.0

        dv = _jsonable_scalar(_cell(padded, ix_date))
        if dv is None or dv == "":
            title_probe = _jsonable_scalar(_cell(padded, ix_title))
            amt_probe = _cell(padded, ix_amount)
            memo_probe = _jsonable_scalar(_cell(padded, ix_memo))
            if all(x in (None, "") for x in (title_probe, amt_probe, memo_probe)):
                continue
            warnings.append(f"행 {excel_row_no}: 날짜 없음 건너뜁니다.")
            continue

        tv = _jsonable_scalar(_cell(padded, ix_title))
        mv = _jsonable_scalar(_cell(padded, ix_memo))
        cv = _jsonable_scalar(_cell(padded, ix_cat))

        data_obj: dict[str, Any] = {
            "date": dv,
            "amount": amt,
            "source": "excel_import",
            "sheet": sheet,
            "excel_row": excel_row_no,
        }
        if tv:
            data_obj["title"] = tv
        if mv:
            data_obj["memo"] = mv
        if cv:
            data_obj["category"] = cv

        row_sql: dict[str, Any] = {
            "led_id": str(led_id),
            "data_type": data_type,
            "data": data_obj,
        }
        if file_id is not None:
            row_sql["file_id"] = str(file_id)

        out.append(row_sql)

    return out, warnings
